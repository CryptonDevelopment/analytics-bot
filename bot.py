import os
import zipfile
import logging
import asyncio
import asyncpg
import io
import pandas as pd
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.types import (
    InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton, FSInputFile, BufferedInputFile, InputFile
)
from openpyxl import Workbook

# Импорты для работы с состояниями (FSM)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
import csv
from dotenv import load_dotenv

load_dotenv()
# ----------------------------
# Конфигурация и глобальные переменные
# ----------------------------

# Токен Telegram-бота (установите свои креды)
# Токен Telegram-бота (установите свои креды)
STATS_DB_URL = os.getenv("STATS_DB_URL", "")
# Токен Telegram-бота
BOT_TOKEN = os.getenv("BOT_TOKEN", "")


# URL для подключения к базам данных по направлениям (установите свои креды)
DATABASE_URLS = {
    "nutsfarm": os.getenv("NUTSFARM_DB_URL", ""),
    "union": os.getenv("UNION_DB_URL", ""),
    "analytics_bot": os.getenv("CHATS_ANALYTICS_DB_URL", ""),
}

# Разрешённые пользователи с указанием отдела.
# Формат переменной ALLOWED_USERS: "123456789:marketing,987654321:analytics,111111111:admin"
allowed_users_env = os.getenv("ALLOWED_USERS", "")
ALLOWED_USERS = {}
for entry in allowed_users_env.split(","):
    parts = entry.split(":")
    if len(parts) == 2:
        try:
            uid = int(parts[0].strip())
            dept = parts[1].strip().lower()
            ALLOWED_USERS[uid] = dept
        except ValueError:
            pass

# Конфигурация сервисов, отображаемых в постоянной клавиатуре.
SERVICES_CONFIG = {
    "marketing": [
        {"name": "Union-M", "id": "union_marketing"},
        {"name": "Calendar-M", "id": "calendar_marketing"},
        {"name": "Nutsfarm-M", "id": "nutsfarm_marketing"},
        {"name": "Tg-chats-M", "id": "telegram_chats"}
    ],
    "analytics": [
        {"name": "Nutsfarm-A", "id": "analytics_nuts"}
    ]
}

# Конфигурация запросов для каждого сервиса.
SERVICE_QUERIES = {
    "telegram_chats": [
        {"name": "Аналитика телеграм чатов", "callback": "qTGa", "sql": "SELECT * FROM activity_stats;", "db": "analytics_bot", "db_instanse": "analytics_bot"}, 
    ],
    "union_marketing": [
       
    ],
    "calendar_marketing": [
       
    ],
    "nutsfarm_marketing": [
        {"name": "DAU, WAU, MAU", "callback": "DAU, WAU, MAU", "sql": """SELECT 
    (SELECT COUNT(*) FROM (SELECT l.user_id FROM user_transaction l WHERE l.created_at >= date_trunc('day', now() - interval '1 day') AND l.created_at < date_trunc('day', now()) GROUP BY l.user_id HAVING COUNT(l.id) >= 1) AS active_users) AS dau,
    (SELECT COUNT(*) FROM (SELECT l.user_id FROM user_transaction l WHERE l.created_at >= now() - interval '1 week' GROUP BY l.user_id HAVING COUNT(l.id) >= 1) AS active_users) AS wau,
    (SELECT COUNT(*) FROM (SELECT l.user_id FROM user_transaction l WHERE l.created_at >= now() - interval '30 day' GROUP BY l.user_id HAVING COUNT(l.id) > 1) AS active_users) AS mau""", "db": "nutsfarm", "db_instanse": "nutsfarm"},
    {"name": "RET_1d, RET_3d, RET_7d, RET_30d", "callback": "RET_1d, RET_3d, RET_7d, RET_30d", "sql": """SELECT
    (SELECT COUNT(DISTINCT user_id) FROM user_transaction l WHERE l.created_at >= date_trunc('day', now() - interval '1 day') AND l.created_at < date_trunc('day', now())) AS ret_1d,
    (SELECT COUNT(DISTINCT user_id) FROM user_transaction l WHERE l.created_at >= date_trunc('day', now() - interval '3 day') AND l.created_at < date_trunc('day', now())) AS ret_3d,
    (SELECT COUNT(DISTINCT user_id) FROM user_transaction l WHERE l.created_at >= date_trunc('day', now() - interval '7 day') AND l.created_at < date_trunc('day', now())) AS ret_7d,
    (SELECT COUNT(DISTINCT user_id) FROM user_transaction l WHERE l.created_at >= date_trunc('day', now() - interval '30 day') AND l.created_at < date_trunc('day', now())) AS ret_30d
""", "db": "nutsfarm", "db_instanse": "nutsfarm"},    {"name": "New Users (первичный вход)", "callback": "New Users", "sql": """SELECT COUNT(*) AS new_users
FROM entity_user e
WHERE e.created_at >= date_trunc('day', now() - interval '1 day')
""", "db": "nutsfarm", "db_instanse": "nutsfarm"},     {"name": "Revenue, ARPU, ARPPU", "callback": "Revenue, ARPU, ARPPU", "sql": """SELECT 
    (SELECT SUM(amount) FROM payment_transactions WHERE created_at >= now() - interval '1 month') AS revenue,
    (SELECT SUM(amount) / COUNT(DISTINCT user_id) FROM payment_transactions WHERE created_at >= now() - interval '1 month') AS arpu,
    (SELECT SUM(amount) / COUNT(DISTINCT CASE WHEN amount > 0 THEN user_id END) FROM payment_transactions WHERE created_at >= now() - interval '1 month') AS arppu
""", "db": "nutsfarm", "db_instanse": "nutsfarm"},      {"name": "Churn Rate (_1d, _3d, _7d, _30d)", "callback": "Churn Rate (_1d, _3d, _7d, _30d)", "sql": """WITH user_activity AS (
    SELECT
        user_id,
        MAX(CASE WHEN created_at >= date_trunc('day', now() - interval '1 day') THEN 1 ELSE 0 END) AS active_1d,
        MAX(CASE WHEN created_at >= date_trunc('day', now() - interval '3 day') THEN 1 ELSE 0 END) AS active_3d,
        MAX(CASE WHEN created_at >= date_trunc('day', now() - interval '7 day') THEN 1 ELSE 0 END) AS active_7d,
        MAX(CASE WHEN created_at >= date_trunc('day', now() - interval '30 day') THEN 1 ELSE 0 END) AS active_30d
    FROM
        user_transaction
    GROUP BY
        user_id
)
SELECT
    -- Чурн по 1 дню
    COUNT(CASE WHEN active_1d = 0 THEN 1 END) * 1.0 / COUNT(user_id) AS churn_1d,

    -- Чурн по 3 дням
    COUNT(CASE WHEN active_3d = 0 THEN 1 END) * 1.0 / COUNT(user_id) AS churn_3d,

    -- Чурн по 7 дням
    COUNT(CASE WHEN active_7d = 0 THEN 1 END) * 1.0 / COUNT(user_id) AS churn_7d,

    -- Чурн по 30 дням
    COUNT(CASE WHEN active_30d = 0 THEN 1 END) * 1.0 / COUNT(user_id) AS churn_30d
FROM
    user_activity;""", "db": "nutsfarm", "db_instanse": "nutsfarm"},
    ],
    "analytics_nuts": [
        {"name": "Выгрузка всех пользователей", "callback": "qX", "sql": """WITH user_data AS (
        SELECT
            u.id,
            u.telegram_id,
            u.username,
            u.balance,
            COALESCE(ll.lessons_count, 0) AS lessons_count,
            u.referral_count,
            COALESCE(pt.donation_stars, 0) AS donation_stars,
            COALESCE(task.tasks_count, 0) AS tasks_count,
            COALESCE(st.current_streak, 0) AS current_streak,
            CASE
                WHEN u.ton_wallet IS NOT NULL AND TRIM(u.ton_wallet) <> '' THEN true
                ELSE false
            END AS has_ton_wallet
        FROM entity_user u
        LEFT JOIN (
            SELECT user_id, COUNT(*) AS lessons_count
            FROM link_user_learn_lesson
            GROUP BY user_id
        ) ll ON ll.user_id = u.id
        LEFT JOIN (
            SELECT user_id, COUNT(*) AS tasks_count
            FROM link_user_task
            WHERE status = 'CLAIMED'
            GROUP BY user_id
        ) task ON task.user_id = u.id
        LEFT JOIN (
            SELECT user_id, COALESCE(SUM(amount), 0) AS donation_stars
            FROM payment_transactions
            WHERE status = 'SUCCESS'
            GROUP BY user_id
        ) pt ON pt.user_id = u.id
        LEFT JOIN (
            SELECT DISTINCT ON (user_id) user_id, current_streak
            FROM link_user_streak_day
            ORDER BY user_id, updated_at DESC
        ) st ON st.user_id = u.id
    )
    SELECT
        id,
        telegram_id,
        username,
        balance,
        lessons_count,
        referral_count,
        donation_stars,
        tasks_count,
        current_streak,
        has_ton_wallet
    FROM user_data;""", "db": "nutsfarm", "db_instanse": "nutsfarm"},
        {"name": "Выгрузка активированности юзеров", "callback": "qW", 
         "sql": """WITH lessons AS (
    SELECT user_id, COUNT(DISTINCT lesson_id) AS lessons_passed
    FROM link_user_learn_lesson
    GROUP BY user_id
),
claims AS (
    SELECT uf.user_id, COUNT(*) AS claims_count
    FROM user_farming_claim ufc
    JOIN user_farming uf ON uf.id = ufc.user_farming_id
    GROUP BY uf.user_id
),
donations AS (
    SELECT user_id, COALESCE(SUM(amount), 0) AS donated_amount
    FROM payment_transactions
    GROUP BY user_id
),
stories AS (
    SELECT user_id, COUNT(*) AS story_views
    FROM link_user_story
    GROUP BY user_id
),
max_nut AS (
    SELECT user_id, MAX(current_streak) AS max_nut_run_day
    FROM link_user_streak_day
    GROUP BY user_id
),
current_nut AS (
    SELECT DISTINCT ON (user_id) user_id, current_streak AS current_nut_run_day
    FROM link_user_streak_day
    ORDER BY user_id, created_at DESC
),
tasks AS (
    SELECT user_id, COUNT(*) AS tasks_completed
    FROM link_user_task
    GROUP BY user_id
)
SELECT
    eu.username AS tg,
    eu.telegram_id AS tg_id,
    eu.crypton_global_id AS crypton_id,
    eu.balance,
    COALESCE(l.lessons_passed, 0) AS lessons_passed,
    eu.referral_count,
    (CURRENT_DATE - DATE(eu.created_at)) AS days_in_app,
    COALESCE(c.claims_count, 0) AS claims_count,
    COALESCE(d.donated_amount, 0) AS donated_amount,
    COALESCE(s.story_views, 0) AS story_views,
    COALESCE(m.max_nut_run_day, 0) AS max_nut_run_day,
    COALESCE(cn.current_nut_run_day, 0) AS current_nut_run_day,
    COALESCE(t.tasks_completed, 0) AS tasks_completed
FROM entity_user eu
LEFT JOIN lessons l ON l.user_id = eu.id
LEFT JOIN claims c ON c.user_id = eu.id
LEFT JOIN donations d ON d.user_id = eu.id
LEFT JOIN stories s ON s.user_id = eu.id
LEFT JOIN max_nut m ON m.user_id = eu.id
LEFT JOIN current_nut cn ON cn.user_id = eu.id
LEFT JOIN tasks t ON t.user_id = eu.id
ORDER BY eu.id;
""", 
         "db": "nutsfarm", "db_instanse": "nutsfarm"}
    ]
}

# Глобальная переменная для отслеживания активности пользователей в приватных чатах.
ACTIVITY_STATS = {}

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ----------------------------
# Модуль миграций
# ----------------------------

def generate_csv(data) -> io.BytesIO:
    output = io.StringIO()
    if data:
        # Преобразуем первую запись в dict, чтобы получить заголовки
        headers = list(dict(data[0]).keys())
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for record in data:
            # Преобразуем каждую запись в dict перед записью
            writer.writerow(dict(record))
    else:
        output.write("Нет данных для отображения")
    csv_bytes = io.BytesIO(output.getvalue().encode("utf-8"))
    csv_bytes.seek(0)
    return csv_bytes

def compress_excel_to_zip(excel_bytes_io: io.BytesIO, zip_filename: str = "report.zip") -> io.BytesIO:
    # Получаем имя Excel файла, если оно установлено, иначе используем "report.xlsx"
    excel_filename = getattr(excel_bytes_io, "name", "report.xlsx")
    zip_bytes_io = io.BytesIO()
    # Используем метод ZIP_LZMA с максимальным уровнем сжатия (compresslevel=9)
    with zipfile.ZipFile(zip_bytes_io, mode="w", compression=zipfile.ZIP_LZMA, compresslevel=9) as zf:
        zf.writestr(excel_filename, excel_bytes_io.getvalue())
    zip_bytes_io.seek(0)
    return zip_bytes_io

async def init_stats_table():
    """
    Инициализация таблицы для статистики сообщений.
    Таблица создаётся с дополнительными полями для названия чата и топика (если он есть).
    """
    conn = await asyncpg.connect(STATS_DB_URL)
    try:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS activity_stats (
                chat_id BIGINT,
                user_id BIGINT,
                chat_title TEXT,
                chat_topic TEXT,
                message_count INTEGER,
                total_length INTEGER,
                PRIMARY KEY (chat_id, user_id)
            );
        """)
        logger.info("Таблица activity_stats успешно инициализирована.")
    finally:
        await conn.close()

async def run_migrations():
    """
    Запускает все миграции для БД.
    Здесь можно добавить и другие миграционные шаги.
    """
    await init_stats_table()
    logger.info("Все миграции успешно выполнены.")

# ----------------------------
# Вспомогательные функции
# ----------------------------

async def fetch_data(query: str, db_url: str):
    conn = await asyncpg.connect(db_url)
    try:
        data = await conn.fetch(query)
    finally:
        await conn.close()
    return data

def generate_excel(data) -> io.BytesIO:
    max_rows = 1048576
    wb = Workbook(write_only=True)

    if not data:
        ws = wb.create_sheet("Sheet1")
        ws.append(["Нет данных для отображения"])
    else:
        headers = list(data[0].keys())
        chunk_size = max_rows - 1  
        for i in range(0, len(data), chunk_size):
            sheet_title = f"Sheet{i//chunk_size + 1}"
            ws = wb.create_sheet(title=sheet_title)
            ws.append(headers)
            for record in data[i:i + chunk_size]:
                ws.append([record.get(header) for header in headers])
    
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    return excel_io

def get_user_department(user_id: int):
    return ALLOWED_USERS.get(user_id)

def get_services_for_user(dept: str):
    services = []
    if dept == "admin":
        for _, value in SERVICES_CONFIG.items():
            for elem in value:
                services.append(elem)
    else:
        services = SERVICES_CONFIG.get(dept, [])
    return services

def get_reply_keyboard_for_services(dept: str):
    services = get_services_for_user(dept)
    # Формируем список рядов кнопок – каждый ряд содержит одну кнопку
    buttons = [[KeyboardButton(text=service["name"])] for service in services]
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)

def get_service_by_name(dept: str, name: str):
    services = get_services_for_user(dept)
    for service in services:
        if service["name"] == name:
            return service
    return None

def get_service_group(service_id: str):
    for group, services in SERVICES_CONFIG.items():
        for service in services:
            if service["id"] == service_id:
                return group
    return None

def get_inline_keyboard_for_service(service_id: str):
    queries = SERVICE_QUERIES.get(service_id, [])
    # Создаем InlineKeyboardMarkup с обязательным полем inline_keyboard
    keyboard = InlineKeyboardMarkup(inline_keyboard=[])
    for query in queries:
        callback_data = f"{service_id}:{query['callback']}"
        keyboard.inline_keyboard.append([InlineKeyboardButton(text=query["name"], callback_data=callback_data)])
    return keyboard

# ----------------------------
# Обновление статистики в БД (с chat_title и chat_topic)
# ----------------------------

async def update_activity_stats_in_db(chat_id: int, user_id: int, chat_title: str, chat_topic: str, msg_length: int):
    """
    Обновляет статистику активности пользователя в конкретном чате в БД.
    Если запись существует (по паре chat_id и user_id), то увеличивает счетчик и суммарную длину,
    а также обновляет название чата и топик.
    """
    query = """
    INSERT INTO activity_stats (chat_id, user_id, chat_title, chat_topic, message_count, total_length)
    VALUES ($1, $2, $3, $4, 1, $5)
    ON CONFLICT (chat_id, user_id) DO UPDATE
    SET message_count = activity_stats.message_count + 1,
        total_length = activity_stats.total_length + $5,
        chat_title = EXCLUDED.chat_title,
        chat_topic = EXCLUDED.chat_topic;
    """
    conn = await asyncpg.connect(STATS_DB_URL)
    try:
        await conn.execute(query, chat_id, user_id, chat_title, chat_topic, msg_length)
    finally:
        await conn.close()

# ----------------------------
# FSM для выбора сервиса
# ----------------------------

class ServiceSelection(StatesGroup):
    waiting_for_service = State()

# ----------------------------
# Хэндлеры
# ----------------------------

async def start_handler(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if user_id not in ALLOWED_USERS:
        await message.answer("У вас нет доступа к этому боту.")
        logger.warning("Попытка доступа неавторизованного пользователя: %s", user_id)
        return

    dept = get_user_department(user_id)
    keyboard = get_reply_keyboard_for_services(dept)
    await message.answer("Добро пожаловать!\nВыберите сервис:", reply_markup=keyboard)
    logger.info("Пользователь %s (отдел %s) запустил бота", user_id, dept)
    # Устанавливаем состояние для выбора сервиса
    await state.set_state(ServiceSelection.waiting_for_service)

async def service_selection_handler(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    dept = get_user_department(user_id)
    service = get_service_by_name(dept, message.text.strip())
    if not service:
        await message.answer("Неизвестный сервис. Пожалуйста, выберите сервис с клавиатуры.")
        return
    inline_kb = get_inline_keyboard_for_service(service["id"])
    await message.answer(f"Сервис «{service['name']}». Выберите запрос:", reply_markup=inline_kb)
    logger.info("Пользователь %s выбрал сервис %s", user_id, service["name"])
    # Сбрасываем состояние после выбора сервиса
    await state.clear()

def compress_csv_to_zip(csv_bytes_io: io.BytesIO, zip_filename: str = "report.zip") -> io.BytesIO:
    # Получаем имя CSV-файла, если оно задано, иначе используем дефолтное
    csv_filename = getattr(csv_bytes_io, "name", "report.csv")
    zip_bytes_io = io.BytesIO()
    with zipfile.ZipFile(zip_bytes_io, mode="w", compression=zipfile.ZIP_LZMA, compresslevel=9) as zf:
        zf.writestr(csv_filename, csv_bytes_io.getvalue())
    zip_bytes_io.seek(0)
    # Устанавливаем имя для zip-архива
    zip_bytes_io.name = zip_filename
    return zip_bytes_io

async def query_callback_handler(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    try:
        service_id, query_key = callback.data.split(":", 1)
    except ValueError:
        await callback.answer("Неверные данные запроса.")
        return

    queries = SERVICE_QUERIES.get(service_id, [])
    query_config = None
    for q in queries:
        if q["callback"] == query_key:
            query_config = q
            break

    if not query_config:
        await callback.answer("Запрос не найден.")
        logger.warning("Запрос не найден для callback data: %s", callback.data)
        return

    await callback.answer("Обработка запроса...")

    # Выбор экземпляра БД по параметру из запроса
    db_instance = query_config.get("db_instanse")
    if not db_instance:
        await callback.message.answer("Запрос не настроен: не указан экземпляр БД.")
        logger.error("Нет параметра db_instanse для запроса %s", query_config["name"])
        return

    db_url = DATABASE_URLS.get(db_instance)
    if not db_url:
        await callback.message.answer("Нет настроенной базы данных для данного запроса.")
        logger.error("Нет БД для db_instanse: %s", db_instance)
        return

    try:
        data_records = await fetch_data(query_config["sql"], db_url)
        logger.info("Данные успешно получены для запроса %s сервиса %s пользователем %s",
                    query_config["name"], service_id, user_id)
    except Exception as e:
        logger.error("Ошибка при выполнении запроса: %s", e)
        await callback.message.answer("Ошибка при выполнении запроса к базе данных.")
        return

    await callback.message.answer("Формирую отчет, подождите...")

    try:
        csv_file = generate_csv(data_records)
        csv_file.name = f"{service_id}_{query_key}.csv"
    except Exception as e:
        logger.error("Ошибка при генерации CSV-файла: %s", e)
        await callback.message.answer("Ошибка при генерации CSV файла.")
        return

    # Пороговый размер для сжатия (например, 50 МБ)
    THRESHOLD_SIZE = 50 * 1024 * 1024  
    if len(csv_file.getvalue()) > THRESHOLD_SIZE:
        compressed_file = compress_csv_to_zip(csv_file, zip_filename=f"{service_id}_{query_key}.zip")
        file_to_send = compressed_file
    else:
        file_to_send = csv_file

    try:
        os.makedirs("docs", exist_ok=True)
        file_path = os.path.join("docs", file_to_send.name)
        with open(file_path, "wb") as f:
            f.write(file_to_send.getvalue())
        fs_input_file = FSInputFile(file_path)
        await callback.message.answer_document(
            document=fs_input_file,
            caption=f"Отчет: {query_config['name']}"
        )
        logger.info("CSV-файл отправлен пользователю %s", user_id)
    except Exception as e:
        logger.error("Ошибка при отправке файла: %s", e)
        await callback.message.answer("Ошибка при отправке файла.")
    await state.set_state(ServiceSelection.waiting_for_service)


async def track_activity(message: types.Message):
    if not message.from_user:
        return
    user_id = message.from_user.id
    text = message.text or ""
    stats = ACTIVITY_STATS.get(user_id, {"count": 0, "total_length": 0})
    stats["count"] += 1
    stats["total_length"] += len(text)
    ACTIVITY_STATS[user_id] = stats

    # Если сообщение в групповом чате – обновляем статистику в БД,
    # передавая также название чата и топик (если есть)
    if message.chat.type in ["group", "supergroup"]:
        chat_title = message.chat.title if hasattr(message.chat, "title") else ""
        chat_topic = getattr(message.chat, "topic", None)
        try:
            await update_activity_stats_in_db(message.chat.id, user_id, chat_title, chat_topic, len(text))
        except Exception as e:
            logger.error("Ошибка при обновлении статистики в БД: %s", e)

async def my_stats_handler(message: types.Message):
    user_id = message.from_user.id
    stats = ACTIVITY_STATS.get(user_id, {"count": 0, "total_length": 0})
    count = stats["count"]
    avg_length = stats["total_length"] / count if count > 0 else 0
    await message.answer(f"Ваша активность:\nСообщений: {count}\nСредняя длина: {avg_length:.2f} символов")

async def all_stats_handler(message: types.Message):
    user_id = message.from_user.id
    if get_user_department(user_id) != "admin":
        await message.answer("У вас нет доступа к этой команде.")
        return

    if not ACTIVITY_STATS:
        await message.answer("Нет данных по активности.")
        return

    lines = []
    for uid, stats in ACTIVITY_STATS.items():
        count = stats["count"]
        avg_length = stats["total_length"] / count if count > 0 else 0
        lines.append(f"User {uid}: сообщений {count}, ср. длина {avg_length:.2f}")
    await message.answer("\n".join(lines))

# ----------------------------
# Регистрация хэндлеров через Dispatcher
# ----------------------------

def register_handlers(dp: Dispatcher):
    dp.message.register(start_handler, Command("start"))
    dp.message.register(service_selection_handler, ServiceSelection.waiting_for_service)
    dp.message.register(track_activity)
    dp.message.register(my_stats_handler, Command("my_stats"))
    dp.message.register(all_stats_handler, Command("all_stats"))
    dp.callback_query.register(query_callback_handler, lambda c: c.data and (":" in c.data))

# ----------------------------
# Основной запуск бота
# ----------------------------

async def main():
    bot = Bot(token=BOT_TOKEN)
    dp = Dispatcher()
    register_handlers(dp)
    # Запуск миграций перед стартом бота
    await run_migrations()
    logger.info("Бот запускается...")
    await dp.start_polling(bot)

if __name__ == '__main__':
    asyncio.run(main())
